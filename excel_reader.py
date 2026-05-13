from openpyxl import load_workbook

import config
from text_utils import normalize


def read_input_path(workbook_path: str) -> str:
    """Lee la ruta de entrada desde INPUT_PATH_SHEET!INPUT_PATH_CELL.

    Devuelve cadena vacia si la hoja/celda no existe o esta vacia; el llamador
    decide el fallback.
    """
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if config.INPUT_PATH_SHEET not in wb.sheetnames:
            return ""
        ws = wb[config.INPUT_PATH_SHEET]
        v = ws[config.INPUT_PATH_CELL].value
        return "" if v is None else str(v).strip()
    finally:
        wb.close()


def read_obras(workbook_path: str) -> list:
    """Lee todas las obras uniendo MASTER_SHEET con cada hoja por tipo en NOMBRE.

    Devuelve una lista de diccionarios por obra:

        {
            "nombre":         "BLANQUEAMIENTO",
            "codigo":         "17810308",
            "fecha_recibido": "29/04/2026",
            "fecha_alta_sgs": "06/05/2026",
            "funcionario":    "HAROLD HERRERA",
            "doc_codes": {
                "CARATULA":    "C050634",
                "FORMATO":     "F124162",
                "CERTIFICADO": "D102852",
                "LETRA":       "L097912",
                "PARTITURA":   "P156072",
            },
        }
    """
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        master_rows = _read_master(wb)

        per_type_codes = {}
        for kind in config.STICKER_TYPES:
            sheet_name = config.SHEET_PER_TYPE[kind]
            per_type_codes[kind] = _read_per_type(wb, sheet_name)

        obras = []
        for key, base in master_rows.items():
            doc_codes = {kind: per_type_codes[kind].get(key) for kind in config.STICKER_TYPES}
            obras.append({**base, "doc_codes": doc_codes})
        return obras
    finally:
        wb.close()


def _read_master(wb) -> dict:
    if config.MASTER_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Master sheet '{config.MASTER_SHEET}' not found. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[config.MASTER_SHEET]
    rows_iter = ws.iter_rows(values_only=True)

    # Saltar filas vacias hasta llegar a la fila de encabezados configurada.
    headers_row = _advance_to(rows_iter, config.HEADER_ROW)

    headers = _index_headers(headers_row)
    last_col = _last_nonempty_col(headers_row)

    col_nombre = headers.get(normalize(config.HEADER_NOMBRE))
    col_codigo = headers.get(normalize(config.HEADER_CODIGO))
    col_compositor = headers.get(normalize(config.HEADER_COMPOSITOR))
    col_recibido = headers.get(normalize(config.HEADER_FECHA_RECIBIDO))
    col_alta_sgs = headers.get(normalize(config.HEADER_FECHA_ALTA_SGS))
    if config.FUNCIONARIO_HEADER:
        col_funcionario = headers.get(normalize(config.FUNCIONARIO_HEADER)) or last_col
    else:
        col_funcionario = last_col

    if col_nombre is None:
        raise ValueError(
            f"Header '{config.HEADER_NOMBRE}' not found on sheet '{config.MASTER_SHEET}'. "
            f"Headers found: {sorted(headers)}"
        )

    out = {}
    for row in rows_iter:
        nombre_raw = _at(row, col_nombre)
        if nombre_raw is None or str(nombre_raw).strip() == "":
            continue
        nombre = str(nombre_raw).strip()
        key = normalize(nombre)
        out[key] = {
            "nombre":         nombre,
            "codigo":         _as_str(_at(row, col_codigo)),
            "author":         _as_str(_at(row, col_compositor)),
            "fecha_recibido": _as_date(_at(row, col_recibido)),
            "fecha_alta_sgs": _as_date(_at(row, col_alta_sgs)),
            "funcionario":    _as_str(_at(row, col_funcionario)),
        }
    return out


def _read_per_type(wb, sheet_name: str) -> dict:
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Per-type sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = _advance_to(rows_iter, config.HEADER_ROW)
    headers = _index_headers(headers_row)

    col_nombre = headers.get(normalize(config.HEADER_NOMBRE))
    col_sticker = headers.get(normalize(config.HEADER_STICKER))
    if col_nombre is None or col_sticker is None:
        raise ValueError(
            f"Sheet '{sheet_name}' is missing '{config.HEADER_NOMBRE}' or "
            f"'{config.HEADER_STICKER}' headers. Headers found: {sorted(headers)}"
        )

    out = {}
    for row in rows_iter:
        nombre_raw = _at(row, col_nombre)
        if nombre_raw is None or str(nombre_raw).strip() == "":
            continue
        key = normalize(str(nombre_raw))
        code = _as_str(_at(row, col_sticker))
        if code:
            out[key] = code
    return out


# --- Utilidades internas ---------------------------------------------------


def _advance_to(rows_iter, target_row: int):
    """Consume el iterador hasta `target_row` y devuelve esa fila."""
    row = None
    for _ in range(target_row):
        row = next(rows_iter)
    return row


def _index_headers(headers_row) -> dict:
    """{encabezado_normalizado: indice_columna_1based}.

    Gana la primera aparicion para evitar sobreescrituras silenciosas cuando
    una hoja tiene encabezados duplicados (por ejemplo, multiples STICKER).
    """
    out = {}
    for col_idx, val in enumerate(headers_row, start=1):
        if val is None:
            continue
        key = normalize(str(val))
        if key not in out:
            out[key] = col_idx
    return out


def _last_nonempty_col(headers_row) -> int:
    last = 0
    for i, v in enumerate(headers_row, start=1):
        if v is not None:
            last = i
    return last


def _at(row, col_idx_1based):
    if col_idx_1based is None or col_idx_1based - 1 >= len(row):
        return None
    return row[col_idx_1based - 1]


def _as_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _as_date(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime(config.DATE_FORMAT)
    return str(v).strip()
