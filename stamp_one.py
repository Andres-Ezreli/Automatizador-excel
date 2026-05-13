import os
import sys
from typing import Any, cast

import config
from openpyxl import load_workbook
import stamper

# Configurar salida para evitar errores con caracteres no ASCII.
stdout_reconfigure = getattr(sys.stdout, "reconfigure", None)
if stdout_reconfigure is not None:
    stdout_reconfigure(errors="replace")
stderr_reconfigure = getattr(sys.stderr, "reconfigure", None)
if stderr_reconfigure is not None:
    stderr_reconfigure(errors="replace")


def main() -> int:
    # Argumentos esperados: script, workbook_path, sticker_type, image_path.
    if len(sys.argv) < 4:
        print("Error: argumentos insuficientes.")
        return 1

    workbook_path = sys.argv[1]
    sticker_type = sys.argv[2]
    image_path = sys.argv[3]

    try:
        wb = load_workbook(workbook_path, data_only=True, read_only=True)
        ws = wb[config.STICKERMAKER_SHEET]

        if sticker_type not in config.SINGLE_STAMP_CELLS:
            print(f"Error: tipo '{sticker_type}' no configurado en config.py")
            return 1

        cells = config.SINGLE_STAMP_CELLS[sticker_type]

        # Leer rutas desde Excel.
        src_pdf = str(ws[cells["input_pdf"]].value or "").strip()
        out_name = str(ws[cells["out_name"]].value or "").strip()
        out_dir = str(ws[cells["out_dir"]].value or "").strip()

        # Extraer coordenadas eje_y y eje_x con valores por defecto si la celda esta vacia.
        eje_y = 0.0
        eje_x = 0.0

        if "eje_y" in cells and ws[cells["eje_y"]].value is not None:
            try:
                eje_y = float(ws[cells["eje_y"]].value)
            except ValueError:
                pass

        if "eje_x" in cells and ws[cells["eje_x"]].value is not None:
            try:
                eje_x = float(ws[cells["eje_x"]].value)
            except ValueError:
                pass

        wb.close()

        # Validaciones de archivos.
        if not os.path.isfile(image_path):
            print(f"Error: no se encontro la imagen del sticker en {image_path}")
            return 1

        if not src_pdf or not os.path.isfile(src_pdf):
            print(f"Error: PDF de origen no encontrado en {src_pdf}")
            return 1

        if not out_name.lower().endswith(".pdf"):
            out_name += ".pdf"

        # Preparar salida.
        os.makedirs(out_dir, exist_ok=True)
        dst_pdf = os.path.join(out_dir, out_name)

        # Ejecutar estampado con eje_y y eje_x.
        stamper_api = cast(Any, stamper)
        stamper_api.stamp_pdf(src_pdf, dst_pdf, image_path, eje_y, eje_x)

        print(f"EXITO: PDF creado en {dst_pdf}")

        # Abrir automaticamente el archivo.
        if os.path.exists(dst_pdf):
            os.startfile(dst_pdf)

        return 0

    except Exception as e:
        print(f"Error fatal: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
